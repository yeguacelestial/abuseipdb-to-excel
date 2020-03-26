import requests
import openpyxl
from json import loads
from optparse import OptionParser


def main():
    get_input = argsOptions()

    options = get_input[0]
    arguments = get_input[1]

    API_KEY = options.API_KEY
    nombre_archivo = options.filename

    ips = read_file(nombre_archivo)

    content_list = []
    content_str = []

    for ip in ips:
        response = get_abuseipdb_response(API_KEY, ip)
        response_dict = loads(response)
        data = response_dict['data']

        ip = data['ipAddress']
        country_code = data['countryCode']
        isp = data['isp']
        domain = data['domain']
        countryName = data['countryName']

        content_list.append([countryName, domain, isp, ip])

        content = f"{len(content_list)})\nPaís: {countryName}\nDominio: {domain}\nISP: {isp}\nIP: {ip}\n\n"
        content_str.append(content)

        print(content)
    
    content_str = "".join(content_str)
    create_ascii_file('output.txt', content_str)
    create_excel_file(nombre_archivo[:-4], content_list)


def argsOptions():
    parser = OptionParser(usage="Uso: abuseipdb.py [opciones]")
    parser.add_option("-a", "--archivo", 
                       dest="filename",
                       help="Indica el nombre del archivo de las IPs.")
    parser.add_option("-k", "--api-key",
                        dest="API_KEY",
                        help="Clave API indicada por la página.")
    parser.add_option("-e", "--excel",
                       dest="excel_existen")
    (options, args) = parser.parse_args()

    return options, args


def get_abuseipdb_response(API_KEY, IPNumber):
    headers = {
        'Key': f'{API_KEY}',
        'Accept': 'application/json',
    }

    params = (
        ('maxAgeInDays', '90'),
        ('verbose', ''),
        ('ipAddress', f'{IPNumber}')
    )

    response = requests.get('https://api.abuseipdb.com/api/v2/check', headers=headers, params=params)

    return response.text


def read_file(file):
    with open(f'{file}', 'r') as f:
        ips_list = [line.rstrip('\n') for line in f]
    
    return ips_list


def create_ascii_file(name, content):
    with open(f'{name}', 'w') as f:
        f.write(content)


def create_excel_file(filename:str, data:list):
    filename = f'{filename}.xlsx'
    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title='AbuseIPDB Info')

    sheet = wb.get_sheet_by_name('AbuseIPDB Info')
    sheet['A1'] = 'PAÍS'
    sheet['B1'] = 'DOMINIO'
    sheet['C1'] = 'ISP'
    sheet['D1'] = 'IP'


    for row in range(sheet.max_row-1, len(data)):
        for column in range(4):
            # PAIS:
            sheet.cell(row=row+2, column=1).value = data[row][0]

            # DOMINIO:
            sheet.cell(row=row+2, column=2).value = data[row][1]

            # ISP:
            sheet.cell(row=row+2, column=3).value = data[row][2]

            # IP:
            sheet.cell(row=row+2, column=4).value = data[row][3]

    wb.save(filename=filename)


if __name__ == '__main__':
    main()